# Python 3.8.8
# openpyxl 3.0.8



from tkinter import *
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import time, re


def open_file_lis():
    fileLisNames[fileLisNumber - 2].set(filedialog.askopenfilename(filetypes=[('xls', '*.xlsx')]))
    if fileLisNames[fileLisNumber - 2].get():
        butAddLis.state(['!disabled'])
        check_but_get_report()


def open_file_inf():
    fileInfNames[fileInfNumber - 2].set(filedialog.askopenfilename(filetypes=[('xls', '*.xlsx')]))
    if fileInfNames[fileInfNumber - 2].get():
        butAddInf.state(['!disabled'])
        check_but_get_report()


def open_prev_report_file():
    prevReportNames[prevReportNumber - 2].set(filedialog.askopenfilename(filetypes=[('xls', '*.xlsx')]))
    if prevReportNames[prevReportNumber - 2].get():
        butAddPrevReport.state(['!disabled'])
        check_but_get_report()

    # prev_report_path.set(filedialog.askopenfilename(filetypes=[('xls', '*.xlsx')]))
    # check_but_get_report()


def add_file_prev_report():
    global prevReportNumber
    global butAddPrevReport
    prevReportNumber += 1
    prevReportNames.append(StringVar())
    ttk.Label(framePrevReport, text="Предыдущий отчет:").grid(column=1, row=prevReportNumber, sticky=E)
    ttk.Entry(framePrevReport, width=70, textvariable=prevReportNames[prevReportNumber - 2]).grid(column=2, row=prevReportNumber, sticky=(W, E))
    ttk.Button(framePrevReport, text="Выбрать", command=open_prev_report_file).grid(column=3, row=prevReportNumber, sticky=W)
    if prevReportNumber == 2:
        butAddPrevReport = ttk.Button(framePrevReport, text="+", command=add_file_prev_report)
        butAddPrevReport.grid(column=4, row=prevReportNumber, sticky=W)
    butAddPrevReport.state(['disabled'])

    for child in framePrevReport.winfo_children():
        child.grid_configure(padx=5, pady=5)


def save_file():
    file = filedialog.asksaveasfile(defaultextension=".xls", initialfile=date + ' Отчет Клиника №1', filetypes=[('xls', '*.xlsx')])
    if file:
        saveReportPath.set(file.name)
        check_but_get_report()


def check_but_get_report():
    if fileLisNames[0].get() and fileInfNames[0].get() and saveReportPath.get() and prevReportNames[0].get():
        but_get_otchet.state(['!disabled'])


def add_file_lis():
    global fileLisNumber
    global butAddLis
    fileLisNumber += 1
    ttk.Label(frameLis, text="Файл ЛИС:").grid(column=1, row=fileLisNumber, sticky=E)
    fileLisNames.append(StringVar())
    ttk.Entry(frameLis, width=50, textvariable=fileLisNames[fileLisNumber - 2]).grid(column=2, row=fileLisNumber, sticky=(W, E))
    ttk.Button(frameLis, text="Выбрать", command=open_file_lis).grid(column=3, row=fileLisNumber, sticky=W)
    if fileLisNumber == 2:
        butAddLis = ttk.Button(frameLis, text="+", command=add_file_lis)
        butAddLis.grid(column=4, row=fileLisNumber, sticky=W)
    butAddLis.state(['disabled'])

    for child in frameLis.winfo_children():
        child.grid_configure(padx=5, pady=5)


def add_file_inf():
    global fileInfNumber
    global butAddInf
    fileInfNumber += 1
    ttk.Label(frameInf, text="Файл ИК:").grid(column=5, row=fileInfNumber, sticky=E)
    fileInfNames.append(StringVar())
    ttk.Entry(frameInf, width=50, textvariable=fileInfNames[fileInfNumber - 2]).grid(column=6, row=fileInfNumber, sticky=(W, E))
    ttk.Button(frameInf, text="Выбрать", command=open_file_inf).grid(column=7, row=fileInfNumber, sticky=W)
    if fileInfNumber == 2:
        butAddInf = ttk.Button(frameInf, text="+", command=add_file_inf)
        butAddInf.grid(column=8, row=fileInfNumber, sticky=W)
    butAddInf.state(['disabled'])

    for child in frameInf.winfo_children():
        child.grid_configure(padx=5, pady=5)


def get_report():
    file_out = openpyxl.Workbook()
    sheet_out = file_out.active

    global statNonResult
    global statDouble
    global statNonDocument
    global statDetected
    ri = 0

    def first_row():
        sheet_out['A1'].value = 'ООО "Клиника№1", г.Обнинск, ИНН 4025426126 КПП 402501001'
        sheet_out['A2'].value = 'дата'
        sheet_out['B2'].value = date

        sheet_out['A3'].value = 'Номер заявки (только цифры)'
        sheet_out['B3'].value = 'Фамилия пациента (обязательно)'
        sheet_out['C3'].value = 'Имя пациента (обязательно)'
        sheet_out['D3'].value = 'Отчество пациента (обязательно)'
        sheet_out['E3'].value = 'Дата рождения'
        sheet_out['F3'].value = 'СНИЛС (обязательно) вводить 11 цифр'
        sheet_out['G3'].value = 'Полис ОМС'
        sheet_out[
            'H3'].value = 'Вид документа (Паспорт гражданина РФ, Свидетельство о рождении, Вид на жительство, Заграничный паспорт, Паспорт иностранного гражданина, Иное)'
        sheet_out['I3'].value = 'Серия документа'
        sheet_out['J3'].value = 'Номер документа'
        sheet_out['K3'].value = 'Телефон'
        sheet_out['L3'].value = 'Электронная почта'
        sheet_out['M3'].value = 'Область'
        sheet_out['N3'].value = 'Район'
        sheet_out['O3'].value = 'Город'
        sheet_out['P3'].value = 'Улица'
        sheet_out['Q3'].value = 'Дом'
        sheet_out['R3'].value = 'Корпус'
        sheet_out['S3'].value = 'Квартира'
        sheet_out['T3'].value = 'Пол (М или Ж)'
        sheet_out['U3'].value = 'Диагноз (можно код по МКБ10)'
        sheet_out['V3'].value = 'Дата заявки'
        sheet_out['W3'].value = 'Дата забора биопробы'
        sheet_out['X3'].value = 'Результат анализа ( 1-Положительный, 0-Отрицательный)'
        sheet_out['Y3'].value = 'Вид анализа ( 1 - ПЦР COVID, 2 - Антитела COVID, IgG, 3 - Антитела COVID, IgM)'
        sheet_out[
            'Z3'].value = 'Количественное значение результата (если не заполнено, то анализ считается качественным)'
        sheet_out['AA3'].value = 'Врач'
        sheet_out['AB3'].value = 'Контактный телефон врача'

        headline = NamedStyle(name="headline")
        headline.font = Font(bold=True, size=11)
        headline.alignment = Alignment(wrap_text=True)
        bd = Side(style='thick', color="000000")
        headline.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        file_out.add_named_style(headline)
        # высота заголовка
        # sheet_out['3'].height = '60'
        sheet_out.row_dimensions[3].height = '60'

        for i in range(1, sheet_out.max_column + 1):
            sheet_out.cell(row=3, column=i).style = 'headline'
            # ширина
            sheet_out.column_dimensions[get_column_letter(i)].width = '20'

    time_1 = time.time()
    # print(time_1)

    first_row()
    # row_out = 3

    # перебор файлов инфоклиники
    for path_inf in fileInfNames:
        file_inf = openpyxl.load_workbook(path_inf.get())
        sheet_inf = file_inf.active

        # перебор сторк инфоклиники
        for row_inf in range(2, sheet_inf.max_row + 1):
            f = False
            # перебор прошлых отчетов
            for path_prev_report in prevReportNames:
                prev_report_file = openpyxl.load_workbook(path_prev_report.get())
                prev_report_sheet = prev_report_file.active
                # перебор строк вчерашнего отчета
                for row_prev in range(2, prev_report_sheet.max_row + 1):
                    if sheet_inf.cell(row=row_inf, column=3).value == prev_report_sheet.cell(row=row_prev, column=1).value:
                        statDouble += 1
                        f = True
                        break
                if f:
                    break
            if f:
                continue

            row_out = sheet_out.max_row + 1

            # перебор файлов лис
            for path_lis in fileLisNames:
                file_lis = openpyxl.load_workbook(path_lis.get())
                sheet_lis = file_lis.active

                # перебор строк лис
                for row_lis in range(4, sheet_lis.max_row + 1):
                    # f = False
                    if sheet_lis.cell(row=row_lis, column=2).value == sheet_inf.cell(row=row_inf, column=3).value:
                        result = sheet_lis.cell(row=row_lis, column=7).value
                        if not result:
                            statNonResult += 1
                            continue

                        # номер
                        sheet_out.cell(row=row_out, column=1, value=sheet_inf.cell(row=row_inf, column=3).value)

                        # фио
                        full_name = sheet_inf.cell(row=row_inf, column=4).value.split(' ')
                        last_name = full_name[0]
                        name = full_name[1]
                        patronymic = ''
                        if len(full_name) > 2:
                            for i in range(2, len(full_name)):
                                patronymic += full_name[i] + ' '

                        sheet_out.cell(row=row_out, column=2, value=last_name)
                        sheet_out.cell(row=row_out, column=3, value=name)
                        sheet_out.cell(row=row_out, column=4, value=patronymic)
                        full_name.clear()

                        # др
                        if sheet_inf.cell(row=row_inf, column=5).value:
                            sheet_out.cell(row=row_out, column=5,
                                           value=sheet_inf.cell(row=row_inf, column=5).value.strftime('%d.%m.%Y'))

                        # снилс
                        sheet_out.cell(row=row_out, column=6, value=sheet_inf.cell(row=row_inf, column=12).value)

                        # вид документа
                        sheet_out.cell(row=row_out, column=8, value=sheet_inf.cell(row=row_inf, column=6).value)

                        # серия
                        if sheet_inf.cell(row=row_inf, column=7).value:
                            document = sheet_inf.cell(row=row_inf, column=7).value.split(' ')
                            nomer = document[-1]
                            seria = ''
                            if len(document) > 1:
                                for i in range(0, len(document) - 1):
                                    seria += document[i]
                            sheet_out.cell(row=row_out, column=9, value=seria)
                            # номер документа
                            sheet_out.cell(row=row_out, column=10, value=nomer)
                            document.clear()
                        else:
                            if ri != row_out:
                                statNonDocument += 1

                        # телефон
                        if sheet_inf.cell(row=row_inf, column=13).value:
                            data_cell = sheet_inf.cell(row=row_inf, column=13).value
                            r = re.findall(r'[9][0-9 .\-\(\)]{8,}[0-9]', str(data_cell))
                            tel = ''
                            if r:
                                for w in r[0]:
                                    if w.isdigit():
                                        tel += w
                                sheet_out.cell(row=row_out, column=11, value=tel)



                        # область
                        sheet_out.cell(row=row_out, column=13, value='Калужская')

                        # пол
                        sheet_out.cell(row=row_out, column=20, value=sheet_inf.cell(row=row_inf, column=8).value)

                        # дата заявки
                        if sheet_inf.cell(row=row_inf, column=9).value:
                            sheet_out.cell(row=row_out, column=22,
                                           value=sheet_inf.cell(row=row_inf, column=9).value.strftime('%d.%m.%Y'))

                        # дата забора
                        if sheet_inf.cell(row=row_inf, column=10).value:
                            sheet_out.cell(row=row_out, column=23,
                                           value=sheet_inf.cell(row=row_inf, column=10).value.strftime('%d.%m.%Y'))
                        # результаты из лис
                        if str(result).upper() == 'ОБНАРУЖЕНО' or str(result).upper() == 'ОБНАРУЖЕНО / DETECTED':
                            sheet_out.cell(row=row_out, column=24, value='1')
                            sheet_out.cell(row=row_out, column=25, value='1')
                            if ri != row_out:
                                statDetected += 1
                            break
                        elif str(result).upper() == 'НЕ ОБНАРУЖЕНО' or str(result).upper() == 'НЕ ОБНАРУЖЕНО / NOT DETECTED':
                            sheet_out.cell(row=row_out, column=24, value='0')
                            sheet_out.cell(row=row_out, column=25, value='1')
                        else:
                            try:
                                float_result = float(result.replace(',', '.'))
                                if float_result > 1.1:
                                    sheet_out.cell(row=row_out, column=24, value='1')
                                    if ri != row_out:
                                        statDetected += 1
                                else:
                                    sheet_out.cell(row=row_out, column=24, value='0')
                                sheet_out.cell(row=row_out, column=25, value='2')
                                sheet_out.cell(row=row_out, column=26, value=result)
                            except AttributeError:
                                continue

                        ri = row_out

    time_2 = time.time()
    # print(time_2)

    file_out.save(saveReportPath.get())

    # print('Время выполнения:', round(time_2 - time_1, 3), 'сек')
    messagebox.showinfo(message='Отчет готов, мэм!\n\n'
                                'Удалены из отчета:\n'
                                '- отправленых ранее: ' + str(statDouble) + ' человек\n'
                                '- без результатов: ' + str(statNonResult) + ' человек\n\n'
                                'Всего в отчете ' + str(sheet_out.max_row-3) + ' человек\n'
                                'Положительных: ' + str(statDetected) + '\n'
                                'Без документов: ' + str(statNonDocument) + '\n\n'
                                'Время выполнения: ' + str(round(time_2 - time_1, 1)) + ' сек')


root = Tk()
root.title("Отчет Клиника №1")

date = datetime.today().strftime('%d.%m.%Y')

statDouble = 0
statNonResult = 0
statNonDocument = 0
statDetected = 0

fileLisNumber = 1
fileLisNames = []
butAddLis = 0

fileInfNumber = 1
fileInfNames = []
butAddInf = 0

saveReportPath = StringVar()

prevReportNumber = 1
prevReportNames = []
butAddPrevReport = 0
prev_report_path = StringVar()

mainframe = ttk.Frame(root, padding="3 3 12 12")
mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

frameLis = ttk.Frame(mainframe, padding='5', relief='solid')
frameLis.grid(column=1, row=1, sticky=(N, W, E, S))

frameInf = ttk.Frame(mainframe, padding='5', relief='solid')
frameInf.grid(column=2, row=1, sticky=(N, W, E, S))

framePrevReport = ttk.Frame(mainframe, padding='5', relief='solid')
framePrevReport.grid(column=1, row=2, columnspan=2, sticky=(N, W, E, S))

frameSave = ttk.Frame(mainframe, padding='5', relief='solid')
frameSave.grid(column=1, row=3, columnspan=2, sticky=(N, W, E, S))

ttk.Label(frameSave, text="Сохранить отчет в:").grid(column=1, row=1, sticky=E)
ttk.Entry(frameSave, width=70, textvariable=saveReportPath).grid(column=2, row=1, sticky=(W, E))
ttk.Button(frameSave, text="Выбрать", command=save_file).grid(column=3, row=1, sticky=E)
but_get_otchet = ttk.Button(frameSave, text=" Создать отчет ", command=get_report)
but_get_otchet.grid(column=4, row=1, sticky=E)
but_get_otchet.state(['disabled'])

add_file_lis()
add_file_inf()
add_file_prev_report()

for child in frameSave.winfo_children():
    child.grid_configure(padx=5, pady=5)

root.mainloop()

